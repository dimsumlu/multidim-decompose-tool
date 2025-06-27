import pandas as pd
import streamlit as st
import numpy as np
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

st.set_page_config(page_title="动态多维拆解分析工具", layout="wide")

st.title("📊 动态多维拆解分析工具")
st.markdown("上传你的 CSV 文件，自动识别维度列，进行结构效应和退费率效应的中心化拆解分析，并支持维度映射、组合标注和带格式导出。")

uploaded_file = st.file_uploader("📁 上传 CSV 文件", type=["csv"])

if uploaded_file:
    df = pd.read_csv(uploaded_file)
    st.success("✅ 文件上传成功！")

    numeric_cols = df.select_dtypes(include=[np.number]).columns.tolist()
    dimension_cols = [col for col in df.columns if col not in numeric_cols]

    if len(numeric_cols) < 4:
        st.error("❌ 数值列少于4列，无法拆解。请确保包含：基期在班、当期在班、基期退费、当期退费。")
    else:
        with st.form("config_form"):
            selected_dims = st.multiselect("🔘 请选择拆解维度列（可多选）", options=dimension_cols, default=dimension_cols)
            mapping_col = st.selectbox("🪄 是否对某个维度进行归类映射？（可选）", options=["不映射"] + selected_dims)
            submit_btn = st.form_submit_button("开始拆解")

        if submit_btn:
            in0, in1, ref0, ref1 = numeric_cols[:4]
            df_copy = df.copy()

            # 应用维度映射方式二
            if mapping_col != "不映射":
                unique_vals = df_copy[mapping_col].dropna().unique().tolist()
                st.subheader(f"🧩 归类映射：{mapping_col}")
                mapping_dict = {}
                for val in unique_vals:
                    new_group = st.selectbox(f"{mapping_col} = {val}", ["小学", "小学（下拉）", "初中", "高中", "未分组"], key=str(val))
                    mapping_dict[val] = new_group
                df_copy[mapping_col] = df_copy[mapping_col].map(mapping_dict)

            grouped = df_copy.groupby(selected_dims, dropna=False).agg({
                in0: 'sum',
                in1: 'sum',
                ref0: 'sum',
                ref1: 'sum'
            }).reset_index()

            # 生成组合标签
            grouped["组合key"] = grouped[selected_dims].astype(str).agg("_".join, axis=1)

            # 构造完整组合集
            base_keys = df_copy.groupby(selected_dims, dropna=False)[in0].sum().reset_index()
            base_keys["组合key"] = base_keys[selected_dims].astype(str).agg("_".join, axis=1)

            curr_keys = df_copy.groupby(selected_dims, dropna=False)[in1].sum().reset_index()
            curr_keys["组合key"] = curr_keys[selected_dims].astype(str).agg("_".join, axis=1)

            all_keys = pd.merge(base_keys[["组合key"]], curr_keys[["组合key"]], how="outer", on="组合key", indicator=True)
            all_keys["组合状态"] = all_keys["_merge"].map({"left_only": "新增组合", "right_only": "消失组合", "both": ""})
            grouped = pd.merge(grouped, all_keys[["组合key", "组合状态"]], on="组合key", how="left")

            in0_ratio_col = in0 + "占比"
            in1_ratio_col = in1 + "占比"
            rate0_col = ref0.replace("人数", "") + "率"
            rate1_col = ref1.replace("人数", "") + "率"

            sum_in0 = grouped[in0].sum()
            sum_in1 = grouped[in1].sum()
            sum_ref0 = grouped[ref0].sum()
            sum_ref1 = grouped[ref1].sum()

            grouped[in0_ratio_col] = grouped[in0] / sum_in0
            grouped[in1_ratio_col] = grouped[in1] / sum_in1
            grouped[rate0_col] = grouped[ref0] / grouped[in0].replace(0, np.nan)
            grouped[rate1_col] = grouped[ref1] / grouped[in1].replace(0, np.nan)

            R0 = sum_ref0 / sum_in0
            grouped["结构效应(pp)"] = (grouped[in1_ratio_col] - grouped[in0_ratio_col]) * (grouped[rate0_col] - R0) * 100
            grouped["退费率效应(pp)"] = grouped[in1_ratio_col] * (grouped[rate1_col] - grouped[rate0_col]) * 100
            grouped["合计影响(pp)"] = grouped["结构效应(pp)"] + grouped["退费率效应(pp)"]

            total_row = pd.DataFrame({
                **{dim: ["总计"] for dim in selected_dims},
                in0: [sum_in0],
                in1: [sum_in1],
                ref0: [sum_ref0],
                ref1: [sum_ref1],
                in0_ratio_col: [1.0],
                in1_ratio_col: [1.0],
                rate0_col: [sum_ref0 / sum_in0],
                rate1_col: [sum_ref1 / sum_in1],
                "结构效应(pp)": [grouped["结构效应(pp)"].sum()],
                "退费率效应(pp)": [grouped["退费率效应(pp)"].sum()],
                "合计影响(pp)": [grouped["合计影响(pp)"].sum()],
                "组合key": ["_总计_"],
                "组合状态": [""]
            })

            result = pd.concat([grouped, total_row], ignore_index=True)

            for col in [in0_ratio_col, in1_ratio_col, rate0_col, rate1_col, "结构效应(pp)", "退费率效应(pp)", "合计影响(pp)"]:
                result[col] = result[col].round(4)

            st.subheader("📄 拆解结果")
            st.dataframe(result.drop(columns=["组合key"]), use_container_width=True)

            # 下载 Excel 带格式
            def to_styled_excel(df):
                wb = Workbook()
                ws = wb.active
                for r in dataframe_to_rows(df.drop(columns=["组合key"]), index=False, header=True):
                    ws.append(r)
                header_font = Font(name='Arial', size=11, bold=True)
                center_align = Alignment(horizontal='center', vertical='center')
                fill = PatternFill("solid", fgColor="DDDDDD")
                for cell in ws[1]:
                    cell.font = header_font
                    cell.alignment = center_align
                    cell.fill = fill
                for row in ws.iter_rows(min_row=2):
                    for cell in row:
                        cell.alignment = center_align
                        if isinstance(cell.value, float) and abs(cell.value) <= 1.0:
                            cell.number_format = '0.00%'

                excel_io = io.BytesIO()
                wb.save(excel_io)
                excel_io.seek(0)
                return excel_io

            excel_data = to_styled_excel(result)
            st.download_button("📥 下载格式化 Excel 文件", data=excel_data, file_name="decomposition_result.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")